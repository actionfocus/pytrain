# -*- coding: utf-8 -*-
"""
Created on Mon Jan 29 21:57:34 2018

@author: actionfocus
"""

# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import wx 
import tushare as ts
import pandas as pd
import datetime
import time
import thread
import openpyxl as pyxl


 
 
class Frame1(wx.Frame):
    
    codelist=[]
    buy_price_list=[]
 
    workBook = pyxl.load_workbook('C:/laptop/00Python/stockList.xlsx')
    
    sheet1 = workBook.get_sheet_by_name('Sheet1')

    for r in range(34): #目前需要根据行数来手动设置
        code=sheet1.cell(row=r+2,column=2)
        codelist.append(str(code.value))
        price=sheet1.cell(row=r+2,column=3)
        buy_price_list.append(price.value)

    var_percent_criteria=5 #定义偏离幅度，需要手工调节
    set_buy_price=pd.Series(buy_price_list,index=codelist)
    stopflag=False
    
    
    def __init__(self,superior):
        wx.Frame.__init__(self,parent=superior,title="Monitor",size=(800,700))
        panel=wx.Panel(self)
        sizer=wx.BoxSizer(wx.VERTICAL)
        self.text1=wx.TextCtrl(panel,value="List:\n",size=(600,600),style=wx.TE_MULTILINE)
        sizer.Add(self.text1,0,wx.ALIGN_TOP | wx.EXPAND)
        button1=wx.Button(panel,label="start")
        sizer.Add(button1)
        self.Bind(wx.EVT_BUTTON,self.OnClick,button1)
        button2=wx.Button(panel,label="stop")
        self.Bind(wx.EVT_BUTTON,self.Stop,button2)
        sizer.Add(button2)
        panel.SetSizerAndFit(sizer)
        panel.Layout()
    
    def monitor(self):
        while(self.stopflag==False):
            self.text1.AppendText("\n----"+self.timestamp()+"----\n")
            index=1
            for code in self.codelist:
                ex_df = self.get_stock_data(code)
                ex_df_new = self.calc_var(ex_df,code)
                if self.check_alert(ex_df_new) is True:
                    self.msg = self.send_msg(ex_df_new,code)
                else:
                    self.msg = str(index)+'  Alert is not triggered on stock '+code+'. '+'Logged on '+ self.timestamp()+' .'
                index = index+1
                self.text1.AppendText(self.msg+"\n")
            del index
            del ex_df
            del ex_df_new
            del code
            for cycle in range(0,180):
                time.sleep(1)
                self.text1.AppendText(".")
            self.text1.AppendText("\n") 
        self.text1.AppendText("Stopped.\n")
        
    def OnClick(self,text):
        
        self.stopflag=False
        thread.start_new_thread(self.monitor,())
        
    def remove_monitor(self):
        self.stopflag=True
        self.text1.Clear()
    
    def Stop(self,text):
        thread.start_new_thread(self.remove_monitor,())
    
    def get_stock_data(self,stockcode):
        df1=ts.get_realtime_quotes(stockcode)
        return df1

    def calc_var(self,df2,stockcode):
        df2.insert(0,'buy_price',self.set_buy_price[stockcode])
        df2.insert(1,'buy_var',(float(df2.price)-df2.buy_price)/float(df2.price)*100)
        return df2
    
    def check_alert(self,df3):
        var_value=df3.buy_var[0]
        criteria_value=self.var_percent_criteria
        if var_value<(0-criteria_value) or (var_value>criteria_value):
            alert=True
        else:
            alert=False
        return alert
    
    def send_msg(self,df4,stockcode):
        info=u'[量化平台消息]'+stockcode+' '+df4.name.iloc[0]+u'变动超出预警值，需关注。'+u'当前价为'+df4.price.iloc[0]+u', 变动比率是'+str(df4.buy_var[0])+'%.'
        return info

    def timestamp(self):
        stamp=datetime.datetime.now().date().strftime('%Y-%m-%d')+'-'+datetime.datetime.now().time().strftime('%H-%M')
        return stamp
    

        
if __name__=="__main__":
    
    app=wx.App()
    frame=Frame1(None)
    frame.Show(True)
    app.MainLoop()
        