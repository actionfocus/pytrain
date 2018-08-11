# -*- coding: utf-8 -*-
"""
Created on Mon Jun 18 22:38:01 2018

@author: actionfocus
"""

import tushare as ts
import openpyxl as pyxl

codelist=[] 
workBook = pyxl.load_workbook('C:/Temp/ETFlist.xlsx')
sheet1 = workBook.get_sheet_by_name('Sheet1')
#mkt_code=[]

wb = pyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title='20DayAvgPrice'

range_item_number=153#目前需要根据行数来手动设置3586,Excel行数-1
start='2018-05-19'
end='2018-06-15'

for r in range(range_item_number): 
   code=sheet1.cell(row=r+2,column=2)
   codelist.append(str(code.value))

count=1
   
for code in codelist:
    if code.startswith('6') or code.startswith('9')\
        or code.startswith('51'): #这个判断规则只适用于A股市场的数据源
        mktcode = 'sh'
    else:
        mktcode = 'sz'
    full_stockcode = mktcode+code
    df=ts.get_k_data(code,start=start,end=end)
    try:
        avg=df['close'].mean()
        flag=0
    except:
        print full_stockcode
        flag=1
    if flag==1:
        dataSet=[full_stockcode,'No data from souce']
    else:
        dataSet=[full_stockcode,float(avg)]
    sheet.append(dataSet)
    print dataSet,',the processed item is ',count
    count=count+1

try:
    wb.save('C:/Temp/ETF_20DayAvg.xlsx')
    print 'Spreadsheet saved.'
except:
    print 'Save failed.'

