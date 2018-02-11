# -*- coding: utf-8 -*-
"""
Created on Sun Feb 11 15:13:32 2018

@author: actionfocus
"""
# -*- coding: utf-8 -*-
#加入上面这一行中文字符处理的错误不会出现

import wx
import examsys
import pandas as pd
import openpyxl as pyxl

class ExamFrame(examsys.MyFrame1):
    
    questionindex=0
    questionDF=pd.DataFrame()
    
    def __init__(self, parent):
        examsys.MyFrame1.__init__(self, parent)
    
    def startTest(self, event):
        #开始测试，显示第一道题目
        self.getStandardQuestion()
        self.m_textCtrl1.SetValue(self.questionDF.iloc[self.questionindex].name)
        #pass
    
    def submitAnswer(self, event):
        #提交当前文本框里的答案
        self.m_textCtrl4.SetValue(str(''))
        result = self.m_textCtrl2.GetValue()
        if self.questionDF.iloc[self.questionindex][0] == str(result):
            self.m_textCtrl3.SetValue(str('答案正确，真棒！'))
        else:
            self.m_textCtrl3.SetValue(str('答案错误，再试试吧！'))
        #pass
    
    def goNextQuestion(self, event):
        #进入下一道题目
        if len(self.questionDF) > self.questionindex+1:
            self.questionindex = self.questionindex + 1
            self.m_textCtrl1.SetValue(self.questionDF.iloc[self.questionindex].name)
            #self.m_textCtrl3.SetValue(str(''))
        else:
            self.m_textCtrl3.SetValue(str('这是最后一题啦。'))
        self.m_textCtrl2.SetValue(str(''))
        self.m_textCtrl4.SetValue(str(''))
        self.m_textCtrl2.SetFocus()
        #pass
    
    def displayAnswer(self, event):
        #显示正确答案
        correctAnswer = self.questionDF.iloc[self.questionindex][0]
        self.m_textCtrl4.SetValue(str(correctAnswer))
        pass
    
    def getStandardQuestion(self):
        #获取标准题库的内容
        questionList = []
        answerList= []
        
        workBook = pyxl.load_workbook('C:/laptop/00Python/dataset/qlib.xlsx')
        sheet1 = workBook.get_sheet_by_name('Sheet1')
        for r in range(50): #目前需要根据行数来手动设置
            question=sheet1.cell(row=r+1,column=2)
            questionList.append(str(question.value))
            answer=sheet1.cell(row=r+1,column=3)
            answerList.append(str(answer.value))       
        self.questionDF = pd.DataFrame(answerList, questionList)
        #pass
    
    def clearJudgement(self, event):
        if len(self.questionDF) > self.questionindex+1:
            self.m_textCtrl3.SetValue(str(''))
    


if __name__=="__main__":
    
    app=wx.App(False)
    frame=ExamFrame(None)
    frame.Show(True)
    app.MainLoop()