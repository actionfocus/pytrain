# -*- coding: utf-8 -*-
"""
Created on Sat Feb 03 20:45:13 2018

@author: 
"""
import pandas as pd
import openpyxl as pyxl

questionList = []
answerList= []
        
workBook = pyxl.load_workbook('C:/laptop/00Python/dataset/qlib.xlsx')
sheet1 = workBook.get_sheet_by_name('Sheet1')
for r in range(7): #目前需要根据行数来手动设置
    question=sheet1.cell(row=r+1,column=2)
    questionList.append(str(question.value))
    answer=sheet1.cell(row=r+1,column=3)
    answerList.append(str(answer.value))
    
df = pd.DataFrame(answerList, questionList)

print df.iloc[1][0]
print len(df)