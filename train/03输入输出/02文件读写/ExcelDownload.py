# -*- coding: utf-8 -*-
"""
Created on Mon Jun 18 09:44:56 2018

@author: actionfocus
"""

import openpyxl as pyxl

#创建工作簿和页签
wb = pyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title='TestSheet'
wb.create_sheet(index=0,title='InsertasFirst')
wb.create_sheet(index=1,title='InsertasSecond')
wb.create_sheet(title='InsertasDefaultLast')


#将值写入Sheet
#wb = pyxl.load_workbook('C:/Temp/pythonDownload.xlsx')
sheet=wb.get_sheet_by_name('InsertasFirst')
sheet['B1']=10
sheet['B2']=20
sheet['B3']=30
sheet['B4']=40
sheet['B5']=50
sheet['A6']='Total'
sheet['B6']='=SUM(B1:B5)'

#保存Excel
try:
    wb.save('C:/Temp/pythonDownload.xlsx')
    print 'Spreadsheet saved.'
except:
    print 'Save failed.'
